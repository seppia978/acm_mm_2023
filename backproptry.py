import os
import numpy as np
import torch
import torch.nn as nn
import torchvision.models as models
import torch.nn.functional as FF
import torchvision.transforms as transforms
from torchvision.datasets import ImageFolder
from torch.utils.data import DataLoader, random_split, Subset
from generics.plotutils import plot_pairs
import pyxlutils

import torchvision.transforms.functional as F
import pandas as pd
import openpyxl as xls
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

img_idx = -1
AH={}
BH={}
def get_layer_act(is_a, name):
    def register_hook(module, inp, out):
        if is_a:
            AH[name] = out[0].requires_grad_()
        else:
            BH[name] = out[0].requires_grad_()
    return register_hook

general = models.vgg16(pretrained=True)

PATH = "/homes/spoppi/pycharm_projects/inspecting_twin_models/checkpoints/short/class_2_untraining_vgg16.pt"
unlearnt = models.vgg16(pretrained=False)
unlearnt.load_state_dict(torch.load(PATH))

wb = Workbook()

ws = wb.active
ws = pyxlutils.create_ws(ws=ws, cname='shark')


c_to_del = 2

out_name = f'activations_{c_to_del}_vgg16.xlsx'

hyperparams = {
    'lr': 1e-3,
    'model': 'resnet18'
}

device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')

size=224
means = [0.485, 0.456, 0.406]
stds = [0.229, 0.224, 0.225]

T = transforms.Compose([
    transforms.Resize(256),
    transforms.CenterCrop(size),
    transforms.ToTensor(),
    transforms.Normalize(means, stds)
])

_T = transforms.Compose([
    transforms.Resize(256),
    transforms.CenterCrop(size),
    transforms.Normalize(means, stds)
])

DeT = transforms.Compose([
    transforms.Normalize(-1 * torch.Tensor(means) / torch.Tensor(stds), 1.0 / torch.Tensor(stds))
])


_train = ImageFolder(
    root='/nas/softechict-nas-2/datasets/Imagenet_new/ILSVRC/Data/CLS-LOC/train',
    transform=T
)

id_to_remove = np.where(np.array(_train.targets) == c_to_del)[0]


train = Subset(_train, id_to_remove)


_val = ImageFolder(
    root='/nas/softechict-nas-2/datasets/Imagenet_new/ILSVRC/Data/CLS-LOC/val',
    transform=T
)

id_to_remove = np.where(np.array(_val.targets) == c_to_del)[0]

val = Subset(_val, id_to_remove)
# test = ImageFolder(
#     root='/nas/softechict-nas-2/datasets/Imagenet_new/ILSVRC/Data/CLS-LOC/test',
#     transform=T
# )


id_others = np.where(np.array(_val.targets) != c_to_del)[0]
others = Subset(_train, id_others)

# loss = nn.CrossEntropyLoss(reduction='none')
# optimizer=torch.optim.SGD(model.parameters(), lr=hyperparams['lr'])

#TEST LOOP
def test_loop(
        n_epochs,
        optimizer,
        model,
        loss_fn,
        train,
        val,
        hyp
):
    global img_idx
    size_others = 10
    others_loader = DataLoader(random_split(others, [size_others, len(others) - size_others])[0], batch_size=64, shuffle=True)

    A,B = model
    A,B=A.cuda(),B.cuda()
    ahook, bhook = {}, {}
    for (an, av), (bn, bv) in zip(A.named_children(), B.named_children()):
        if isinstance(av, nn.Sequential):
            for (an2, av2), (bn2, bv2) in zip(av.named_children(), bv.named_children()):
                if 'basicblock' in an2.lower() or isinstance(av2, models.resnet.BasicBlock):
                    for (an3, av3), (bn3, bv3) in zip(av2.named_children(), bv2.named_children()):
                        ahook[f'{an}_{an2}_{an3}'] = av2.register_forward_hook(get_layer_act(True, f'{an}_{an2}_{an3}'))
                        bhook[f'{bn}_{bn2}_{bn3}'] = bv2.register_forward_hook(get_layer_act(False, f'{bn}_{bn2}_{bn3}'))
                else:
                    ahook[f'{an}_{an2}'] = av2.register_forward_hook(get_layer_act(True, f'{an}_{an2}'))
                    bhook[f'{bn}_{bn2}'] = bv2.register_forward_hook(get_layer_act(False, f'{bn}_{bn2}'))
        else:
            ahook[an] = av.register_forward_hook(get_layer_act(True, an))
            bhook[bn] = bv.register_forward_hook(get_layer_act(False, bn))


    for epoch in range(n_epochs):

        val_loader = DataLoader(random_split(val, [size_others, len(val) - size_others])[0], batch_size=64, shuffle=True)
        img_idx = val_loader.dataset.indices[0]
        print(f'Val {epoch=} ')
        A.train()
        B.train()

        # with torch.no_grad():
        for idx, (imgs, labels) in enumerate(val_loader):
            imgs = imgs.requires_grad_().cuda()
            labels = labels.cuda()

            ascores = nn.functional.softmax(A(imgs),dim=1).sum(dim=0)
            bscores = nn.functional.softmax(B(imgs),dim=1).sum(dim=0)

            A.zero_grad()
            B.zero_grad()

            maxid = bscores.argmax()
            # aback = ascores[:, c_to_del].backward(retain_graph=True)
            ga = torch.autograd.grad(ascores[c_to_del], A.features[28].weight)[0]
            # bback = bscores[:, maxid].backward(retain_graph=True)
            gb = torch.autograd.grad(bscores[maxid], B.features[28].weight)[0]

            plot_pairs(
                ga, gb, 0, 'features.28', idx2=-2,
                path=None
            )


                # cl_acc = (t_out_val.max(1).indices == c_to_del).sum() / t_out_val.max(1).indices.shape[0]

            #     ot_t_acc = []
            # for o, ol in others_loader:
            #     o = o.cuda()
            #     ol = ol.cuda()
            #     test_outs = model(o)
            #     ot_t_acc.append((test_outs.max(1).indices == ol).sum() / (test_outs.max(1).indices == ol).shape[0])


        # if epoch == 1 or epoch % 10 == 0:
        #     print(f"Epoch {epoch}, Training loss {loss_train.item():.4f}, Validation loss {loss_val.item():.4f}")

test_loop(
    n_epochs = 1,
    optimizer = None,
    model = [general, unlearnt],
    loss_fn = None,
    train=train,
    val=val,
    hyp=hyperparams
)

def conv_diff(a,b):
    ret = []
    for i in range(a.shape[0]):
        ret.append([
            [a[i].min(),a[i].max(), a[i].mean(), a[i].var()],
            [b[i].min(),b[i].max(), b[i].mean(), b[i].var()]
        ])

    return ret

def others_diff(a,b):
    ret = []
    ret.append([
        [a.min(),a.max(), a.mean(), a.var()],
        [b.min(),b.max(), b.mean(), b.var()]
    ])

    return ret

def f(a, b, start_row=1,method='diff', N=True):
    global ws

    ret = dict()
    row = start_row
    for i, ((ak, av), (bk, bv)) in enumerate(zip(a.items(), b.items())):
        # print(aa[0], '\t', (aa[1]-bb[1]).sum())
        xa, xb = av.detach().data, bv.detach().data

        if '29' in ak or True:
            try:
                path='/homes/spoppi/pycharm_projects/inspecting_twin_models/activations_png/vgg16'
                path = os.path.join(path, f'class_{c_to_del}',f'imgidx_{img_idx}', ak)
                if not os.path.isdir(path):
                    os.makedirs(path)
                plot_pairs(
                    xa, xb, 0, ak, idx2=-2,
                    path=path
                )
            except Exception as e:
                print(e)


        if 'conv' in ak:
            # xa, xb = norm(xa, xb)
            val = conv_diff(xa[0], xb[0])
        else:
            val = others_diff(xa, xb)

        ws[f'A{row}'] = ak
        row += 1
        first_row = row
        ws, row = pyxlutils.write_layer(
            val=val,
            ws=ws,
            st_row=first_row

        )



f(AH, BH)

ws = pyxlutils.color_scale(ws=ws, cols=['M','N','O','P'])
ws = pyxlutils.adapt_cols_len(ws=ws)
wb.save(out_name)
