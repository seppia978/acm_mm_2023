import os
import numpy as np
import torch
import ast
import torch.nn as nn
import torch.utils.data as data
import torchvision.models as models
import torch.nn.functional as FF
import torchvision.transforms as transforms
from torchvision.datasets import ImageFolder
from torch.utils.data import DataLoader, random_split, Subset, WeightedRandomSampler
from generics.plotutils import plot_pairs
import pyxlutils
from scipy.special import kl_div
from scipy.stats import entropy

from sklearn.feature_selection import mutual_info_classif as MI

import torchvision.transforms.functional as F
import pandas as pd
import openpyxl as xls
from openpyxl import Workbook

from custom_archs import convert_conv2d_to_alpha, get_all_alpha_layers #, invert_alphas
from openpyxl.utils import get_column_letter
import CKA.cka as cka

c_to_del = [2]
immagine, SCORE, CLASSE, LABELS = None, 0, '', []
# which_net = [159, 168, 180, 242, 163, 243, 246, 211, 179, 236]
which_net = [2, 3, 4, 5, 394, 6, 395, 391, 389, 149]
img_idx = -1
AH={}
BH={}
def get_layer_act(is_a, name):
    def register_hook(module, inp, out):
        if is_a:
            AH[name] = out.data
        else:
            BH[name] = out.data
    return register_hook




# PATH = f"/homes/spoppi/pycharm_projects/inspecting_twin_models/checkpoints/short/class_{'-'.join(str(c) for c in c_to_del)}_freeze-before-28_untraining_vgg16.pt"
# PATH = f"/homes/spoppi/pycharm_projects/inspecting_twin_models/checkpoints/short/class_100-classes_freeze-0_untraining_vgg16.pt"
root = '/work/dnai_explainability/unlearning/all_classes'
PATH = f"{root}/resnet18_1.0_100.0_class-{2}_alpha.pt"
# PATH = f"{root}/resnet18_1.0_100.0_class-{c_to_del[0]}_alpha.pt"
net='vgg16' if 'vgg16' in PATH else 'resnet18'

if net == 'resnet18':
    general = models.resnet18(pretrained=True)
    # general = convert_conv2d_to_alpha(models.resnet18(pretrained=False), m=10, standard=False)
    unlearnt = convert_conv2d_to_alpha(models.resnet18(pretrained=False))
elif net == 'vgg16':
    general = convert_conv2d_to_alpha(models.vgg16(pretrained=True), m=10, standard=False)
    unlearnt = convert_conv2d_to_alpha(models.vgg16(pretrained=False))



unlearnt.load_state_dict(torch.load(PATH))
PATH = f"{root}/resnet18_1.0_1000.0_class-76_alpha.pt"
# general.load_state_dict(torch.load(PATH))
au, ag = get_all_alpha_layers(unlearnt), get_all_alpha_layers(general)
matrix = cka.CKA(au, ag)
# general.load_state_dict(torch.load(PATH))
# invert_alphas(unlearnt)

wb = Workbook()

ws = wb.active
ws = pyxlutils.create_ws(ws=ws, cname='shark')




out_name = f'activations_{"-".join(str(c) for c in c_to_del)}_vgg16.xlsx'

hyperparams = {
    'lr': 1e-3,
    'model': 'resnet18'
}

device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')

size=224
means = [0.485, 0.456, 0.406]
stds = [0.229, 0.224, 0.225]

T = transforms.Compose([
    transforms.Resize(256),
    transforms.CenterCrop(size),
    transforms.ToTensor(),
    transforms.Normalize(means, stds)
])

_T = transforms.Compose([
    transforms.Resize(256),
    transforms.CenterCrop(size),
    transforms.Normalize(means, stds)
])

DeT = transforms.Compose([
    transforms.Normalize(-1 * torch.Tensor(means) / torch.Tensor(stds), 1.0 / torch.Tensor(stds))
])


# _train = ImageFolder(
#     root='/nas/softechict-nas-2/datasets/Imagenet_new/ILSVRC/Data/CLS-LOC/train',
#     transform=T
# )
#
# id_to_remove = np.where(np.isin(np.array(_train.targets), c_to_del))[0]
#
#
# train = Subset(_train, id_to_remove)
# train = torch.load(os.path.join('dataset_utils','files','train.pt'))


_val = ImageFolder(
    root='/nas/softechict-nas-2/datasets/Imagenet_new/ILSVRC/Data/CLS-LOC/val',
    transform=T
)
#
# id_to_remove = np.where(np.isin(np.array(_val.targets), c_to_del))[0]
#
# val = Subset(_val, id_to_remove)
# _val = torch.load(os.path.join('dataset_utils','files','val.pt'))
id_c = np.where(np.isin(np.array(_val.targets), c_to_del))[0]
id_others = np.where(~ np.isin(np.array(_val.targets), c_to_del))[0]

val_c = Subset(_val, id_c)
val_c.targets = torch.Tensor(_val.targets).int()[id_c].tolist()
val_others = Subset(_val, id_others)
val_others.targets = torch.Tensor(_val.targets).int()[id_others].tolist()

concat_train = data.ConcatDataset((val_c,val_others))
concat_train.targets = torch.Tensor(_val.targets).int()[id_c].tolist()
concat_train.targets = [*val_c.targets, *val_others.targets]

# perm_c = torch.randperm(torch.Tensor(id_c).shape[0])
# idx_c = perm_c[0]
# idx = id_c[idx_c]
#
# perm_ot = torch.randperm(torch.Tensor(id_others).shape[0])
# idx_ot = perm_ot[:3]
# idx.extend(id_others[idx_ot])

# val = Subset(_val.dataset, id_c)

# val_c = Subset(_val.dataset, id_c)
# val_ot = Subset(_val.dataset, id_others)
# val = (val_c,val_ot)

# test = ImageFolder(
#     root='/nas/softechict-nas-2/datasets/Imagenet_new/ILSVRC/Data/CLS-LOC/test',
#     transform=T
# )

#
with open('class_names/names.txt',  'r') as f:
    txt = f.read()

classes = ast.literal_eval(txt)
concat_train.names=classes

# id_others = np.where(~ np.isin(np.array(_val.targets), c_to_del))[0]
# others = Subset(_train, id_others)
others = torch.load(os.path.join('dataset_utils','files','others.pt'))

# loss = nn.CrossEntropyLoss(reduction='none')
# optimizer=torch.optim.SGD(model.parameters(), lr=hyperparams['lr'])

def up_trim_mean(a:torch.Tensor):
    return (a.sum() - a.max()) / (len(a) - 1)

def down_trim_mean(a:torch.Tensor):
    return (a.sum() - a.min()) / (len(a) - 1)

#TRAINING LOOP
def test_loop(
        n_epochs,
        optimizer,
        model,
        loss_fn,
        val,
        hyp
):
    global img_idx,immagine,SCORE, CLASSE, LABELS
    size_others = 64
    # others_loader = DataLoader(random_split(others, [size_others, len(others) - size_others])[0], batch_size=64, shuffle=True)

    A,B = model
    A,B=A.cuda(),B.cuda()
    ahook, bhook = {}, {}

    for (an, av), (bn, bv) in zip(A.named_children(), B.named_children()):
        if isinstance(av, nn.Sequential):
            for (an2, av2), (bn2, bv2) in zip(av.named_children(), bv.named_children()):
                if 'basicblock' in an2.lower() or isinstance(av2, models.resnet.BasicBlock):
                    for (an3, av3), (bn3, bv3) in zip(av2.named_children(), bv2.named_children()):
                        ahook[f'{an}_{an2}_{an3}'] = av2.register_forward_hook(get_layer_act(True, f'{an}_{an2}_{an3}'))
                        bhook[f'{bn}_{bn2}_{bn3}'] = bv2.register_forward_hook(get_layer_act(False, f'{bn}_{bn2}_{bn3}'))
                else:
                    ahook[f'{an}_{an2}'] = av2.register_forward_hook(get_layer_act(True, f'{an}_{an2}'))
                    bhook[f'{bn}_{bn2}'] = bv2.register_forward_hook(get_layer_act(False, f'{bn}_{bn2}'))
        else:
            ahook[an] = av.register_forward_hook(get_layer_act(True, an))
            bhook[bn] = bv.register_forward_hook(get_layer_act(False, bn))

    y_train = val.targets
    weight = 1. / torch.Tensor([.001 for _ in range(1000)])

    weight[~np.isin(list(range(len(weight))), np.array(c_to_del))] = .5 / (len(weight) - len(c_to_del))
    weight[np.array(c_to_del)] = .5 / len(c_to_del)

    samples_weight = np.array([weight[t] for t in y_train])
    samples_weight = torch.from_numpy(samples_weight)
    sampler = WeightedRandomSampler(samples_weight.type('torch.DoubleTensor'), len(samples_weight))

    val_loader = DataLoader(val, sampler=sampler, batch_size=size_others, num_workers=4)

    for epoch in range(n_epochs):

        # val_loader = DataLoader(Subset(val,[13]), batch_size=64, shuffle=True)

        ##########################################################################


        ##########################################################################

        # val_loader = DataLoader(random_split(val, [size_others, len(val) - size_others])[0], batch_size=64, shuffle=True)

        # img_idx = val_loader.dataset.indices
        print(f'Val {epoch=} ')
        A.eval()
        B.eval()

        with torch.no_grad():
            for idx, (imgs, labels) in enumerate(val_loader):
                immagine = DeT(imgs)
                imgs = imgs.cuda()
                LABELS = labels.tolist()
                labels = labels.cuda()

                ascores = A(imgs)
                bscores = B(imgs)
                SCORE=ascores
                # CLASSE = val_loader.dataset.dataset.names[int(bscores.argmax())]
                aent = torch.Tensor([entropy(nn.functional.softmax(ascores, dim=1).cpu().numpy()[i]) for i in range(len(ascores))])
                bent = torch.Tensor([entropy(nn.functional.softmax(bscores, dim=1).cpu().numpy()[i]) for i in range(len(bscores))])
                amean = aent.mean()
                bmean = bent.mean()

                atmean = up_trim_mean(aent)
                btmean = down_trim_mean(bent)

                print(ascores.topk(10).indices)
                print(bscores.topk(10).indices)

                print('Entropy A:', atmean, aent.min(), aent.max(), aent.std())
                print('Entropy B:', btmean, bent.min(), bent.max(), bent.std())
                break
                x=0

                # cl_acc = (t_out_val.max(1).indices == c_to_del).sum() / t_out_val.max(1).indices.shape[0]

            #     ot_t_acc = []
            # for o, ol in others_loader:
            #     o = o.cuda()
            #     ol = ol.cuda()
            #     test_outs = model(o)
            #     ot_t_acc.append((test_outs.max(1).indices == ol).sum() / (test_outs.max(1).indices == ol).shape[0])


        # if epoch == 1 or epoch % 10 == 0:
        #     print(f"Epoch {epoch}, Training loss {loss_train.item():.4f}, Validation loss {loss_val.item():.4f}")



def conv_diff(a,b):
    ret = []
    for i in range(a.shape[0]):
        ret.append([
            [a[i].min(),a[i].max(), a[i].mean(), a[i].var()],
            [b[i].min(),b[i].max(), b[i].mean(), b[i].var()]
        ])

        # kl1 = kl_div(a[i], b[i])
        # kl2 = kl_div(b[i], a[i])
        #
        # ret.append([
        #     [kl1.min(), kl1.max(), kl1.mean(), kl1.var()],
        #     [kl2.min(), kl2.max(), kl2.mean(), kl2.var()]
        # ])

    return ret

def others_diff(a,b):
    ret = []
    ret.append([
        [a.min(),a.max(), a.mean(), a.var()],
        [b.min(),b.max(), b.mean(), b.var()]
    ])

    # kl1 = kl_div(a, b)
    # kl2 = kl_div(b, a)
    #
    # ret.append([
    #     [kl1.min(), kl1.max(), kl1.mean(), kl1.var()],
    #     [kl2.min(), kl2.max(), kl2.mean(), kl2.var()]
    # ])

    return ret

def f(a, b, start_row=1,method='diff', N=True):
    global ws, MIL

    xa, xb = a['layer4_1_conv2'].detach().data, b['layer4_1_conv2'].detach().data
    xa = xa.cpu() #torch.nn.functional.relu(xa)

    import torch.nn.functional as FF
    import matplotlib.pyplot as plt
    from CKA import cka

    # cka.distance(torch.Tensor([[1,2],[3,4]]).numpy(),torch.Tensor([[4,3],[2,1]]).numpy())
    # cka.distance(torch.sigmoid(alpha).cpu().detach().numpy(), torch.sigmoid(alpha).cpu().detach().numpy())
    # cka.distance(torch.sigmoid(alpha).cpu().detach().numpy(), torch.sigmoid(1-alpha).cpu().detach().numpy())

    plt.imshow(immagine[0].detach().cpu().permute(1, 2, 0).numpy())
    plt.show()

    upsampled_b = FF.interpolate(xb, (224, 224), mode='bilinear', align_corners=False)
    upsampled_a = FF.interpolate(xa, (224, 224), mode='bilinear', align_corners=False).to('cuda')
    au = torch.sigmoid(unlearnt.layer4[1].conv1.alpha)
    # MI(xa[:, :].mean(dim=(-1, -2)).detach().cpu().numpy(), np.ones(10), random_state=0)
    # MI(xa[:, :].mean(dim=(-1, -2)).detach().cpu().numpy(), np.ones(10), random_state=0)
    # MI(xa[:, :].mean(dim=(-1, -2)).detach().cpu().numpy(), np.ones(10), random_state=0)
    # MI(xa[:, :].mean(dim=(-1, -2)).detach().cpu().numpy(), np.ones(10), random_state=0)
    # au, ag = torch.sigmoid(unlearnt.layer4[1].conv1.alpha), torch.sigmoid(general.layer4[1].conv1.alpha)
    # zu = torch.ones([1, 1, 224, 224], device='cuda')
    # for i in torch.where((1 - au) >.6)[0]:
    #     zu *= upsampled_b[:, i]
    # zu_ = zu - zu.min()
    # zu_ = zu / (zu.max()+1e-10)
    # zg = torch.ones([1, 1, 224, 224], device='cuda')
    # for i in torch.where((1 - ag) >.6)[0]:
    #     zg *= upsampled_a[:, i]
    # zg_ = zg - zg.min()
    # zg_ = zg / (zg.max()+1e-10)
    xa_ = xa - xa.min()
    xa_ = xa / xa.max()
    nau=1-au
    mi = MI(xa_[:, nau.topk(10).indices].mean(dim=(-1, -2)).detach().cpu().numpy(),
            np.array([c == 2 for c in LABELS]), random_state=0)
    MIL.append(mi)
    x=0
    xb = xb.cpu() #torch.nn.functional.relu(xb)

    alpha= unlearnt.layer4[1].conv1.alpha # unlearnt.features[28].alpha
    # alpha = torch.sigmoid(alpha)

    if False:
        for i in (alpha).topk(10).indices:
            # test = (upsampled_b[:, i])[(None,) + (...,)] * _T(immagine).cuda()
            # test -= test.min()
            # test /= (test.max() + 1e-12)
            # plt.imshow(test[0].permute(1, 2, 0).detach().cpu().numpy(), cmap='jet')
            # plt.show()

            test = (upsampled_b[:, i].clone())
            # print(test.max(), test.min(), '----')
            # test -= test.min()
            # test /= (test.max() + 1e-12)
            # print(test.max(), test.min())
            plt.imshow(test[0].detach().cpu().numpy())
            plt.show()

        for i in (1-alpha).topk(10).indices:
            test = (upsampled_b[:, i])[(None,) + (...,)] * _T(immagine).cuda()
            test -= test.min()
            test /= (test.max() + 1e-12)
            plt.imshow(test[0].permute(1, 2, 0).detach().cpu().numpy(), cmap='jet')
            plt.show()

            test = (upsampled_b[:, i].clone())
            print(test.max(), test.min(),'----')
            test -= test.min()
            test /= (test.max() + 1e-12)
            print(test.max(), test.min())
            plt.imshow(test[0].detach().cpu().numpy(), cmap='jet')
            plt.show()

        upsum = upsampled_b.sum(dim=1)
        upsum -= upsum.min()
        upsum /= (upsum.max() + 1e-12)
        test = (upsum) * _T(immagine).cuda()
        test -= test.min()
        test /= (test.max() + 1e-12)
        plt.imshow(test[0].permute(1, 2, 0).detach().cpu().numpy(), cmap='jet')
        plt.show()


    # if '29' in ak or True:
    #     try:
    #         path='/homes/spoppi/pycharm_projects/inspecting_twin_models/activations_png/vgg16'
    #         path = os.path.join(path, f'class_{c_to_del}',f'imgidx_{img_idx}', ak)
    #         if not os.path.isdir(path):
    #             os.makedirs(path)
    #         plot_pairs(
    #             xa, xb, 0, ak, idx2=-2,
    #             path=path
    #         )
    #     except Exception as e:
    #         print(e)


    # if 'conv' in ak:
    #     # xa, xb = norm(xa, xb)
    #     val = conv_diff(xa[0], xb[0])
    # else:
    #     val = others_diff(xa, xb)

    # ws[f'A{row}'] = ak
    # row += 1
    # first_row = row
    # ws, row = pyxlutils.write_layer(
    #     val=val,
    #     ws=ws,
    #     st_row=first_row
    #
    # )

MIL = []
for i in range(5):
    test_loop(
        n_epochs = 1,
        optimizer = None,
        model = [general, unlearnt],
        loss_fn = None,
        val=concat_train,
        hyp=hyperparams
    )

    f(AH, BH)

ws = pyxlutils.color_scale(ws=ws, cols=['M','N','O','P'])
ws = pyxlutils.adapt_cols_len(ws=ws)
wb.save(out_name)
