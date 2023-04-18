import os
import numpy as np
import torch
import ast
import torch.nn as nn
import torchvision.datasets
import torchvision.models as models
import torch.nn.functional as FF
import torchvision.transforms as transforms
from torchvision.datasets import ImageFolder
from torch.utils.data import DataLoader, random_split, Subset
from generics.plotutils import plot_pairs
import pyxlutils
from scipy.special import kl_div
from scipy.stats import entropy

import torchvision.transforms.functional as F
import pandas as pd
import openpyxl as xls
from openpyxl import Workbook

from custom_archs import convert_conv2d_to_alpha,invert_alphas
from openpyxl.utils import get_column_letter

c_to_del = [2]
immagine, SCORE, CLASSE = None, 0, ''
# which_net = [159, 168, 180, 242, 163, 243, 246, 211, 179, 236]
which_net = [2, 3, 4, 5, 394, 6, 395, 391, 389, 149]
img_idx = -1
AH={}
BH={}
def get_layer_act(is_a, name):
    def register_hook(module, inp, out):
        if is_a:
            AH[name] = out.data
        else:
            BH[name] = out.data
    return register_hook

general = convert_conv2d_to_alpha(models.resnet18(pretrained=True),m=10)

# PATH = f"/homes/spoppi/pycharm_projects/inspecting_twin_models/checkpoints/short/class_{'-'.join(str(c) for c in c_to_del)}_freeze-before-28_untraining_vgg16.pt"
# PATH = f"/homes/spoppi/pycharm_projects/inspecting_twin_models/checkpoints/short/class_100-classes_freeze-0_untraining_vgg16.pt"
PATH = f"/homes/spoppi/pycharm_projects/inspecting_twin_models/checkpoints/short/resnet18_0.9_1000.0_alpha.pt"
unlearnt = convert_conv2d_to_alpha(models.resnet18(pretrained=False))
unlearnt.load_state_dict(torch.load(PATH))
invert_alphas(unlearnt)

wb = Workbook()

ws = wb.active
ws = pyxlutils.create_ws(ws=ws, cname='shark')




out_name = f'activations_{"-".join(str(c) for c in c_to_del)}_vgg16.xlsx'

hyperparams = {
    'lr': 1e-3,
    'model': 'resnet18'
}

device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')

size=224
means = [0.485, 0.456, 0.406]
stds = [0.229, 0.224, 0.225]

T = transforms.Compose([
    transforms.Resize(256),
    transforms.CenterCrop(size),
    transforms.ToTensor(),
    transforms.Normalize(means, stds)
])

_T = transforms.Compose([
    transforms.Resize(256),
    transforms.CenterCrop(size),
    transforms.Normalize(means, stds)
])

DeT = transforms.Compose([
    transforms.Normalize(-1 * torch.Tensor(means) / torch.Tensor(stds), 1.0 / torch.Tensor(stds))
])

def make_image_4_pointing_game(v:torch.utils.data.Subset, idx_c) -> torch.Tensor :
    size = val.dataset[0][0].shape
    img = torch.zeros(size).unsqueeze(0)
    R,C = pow(len(v.indices),.5), pow(len(v.indices),.5)

    for i, idx in enumerate(v.indices):
        im = v.dataset[idx][0].unsqueeze(0)
        im = FF.interpolate(im, [int(x/(R)) for x in img.shape[-2:]], mode='bilinear', align_corners=False)
        r,c = i // R, i % C
        if idx == idx_c:
            rout,cout = int(r), int(c)
        img[:,:,
            int(r/R * img.shape[-1]) : int(r / R * img.shape[-1] + img.shape[-2] / R),
            int(c * (img.shape[-2]) // C) : int(c * (img.shape[-2]) // C + img.shape[-1] / C)
        ] = im

    return img, (rout,cout)
# _train = ImageFolder(
#     root='/nas/softechict-nas-2/datasets/Imagenet_new/ILSVRC/Data/CLS-LOC/train',
#     transform=T
# )
#
# id_to_remove = np.where(np.isin(np.array(_train.targets), c_to_del))[0]
#
#
# train = Subset(_train, id_to_remove)
# train = torch.load(os.path.join('dataset_utils','files','train.pt'))


_val = ImageFolder(
    root='/nas/softechict-nas-2/datasets/Imagenet_new/ILSVRC/Data/CLS-LOC/val',
    transform=T
)
#
# id_to_remove = np.where(np.isin(np.array(_val.targets), c_to_del))[0]
#
# val = Subset(_val, id_to_remove)
# idx.extend(id_others[idx_ot])
# _val = torch.load(os.path.join('dataset_utils','files','val.pt'))
id_c = np.where(np.isin(np.array(_val.targets), c_to_del))[0]
id_others = np.where(~ np.isin(np.array(_val.targets), c_to_del))[0]

perm_c = torch.randperm(torch.Tensor(id_c).shape[0])
idx_c = perm_c[0]
idx = [id_c[idx_c]]

perm_ot = torch.randperm(torch.Tensor(id_others).shape[0])
idx_ot = perm_ot[:3]
idx.extend(list(id_others[idx_ot]))
idx = list(np.random.permutation(idx))

val = Subset(_val, idx)

# test = ImageFolder(
#     root='/nas/softechict-nas-2/datasets/Imagenet_new/ILSVRC/Data/CLS-LOC/test',
#     transform=T
# )

#
img, coords = make_image_4_pointing_game(val, id_c[idx_c])

R,C = pow(len(val.indices),.5), pow(len(val.indices),.5)

with open('class_names/names.txt',  'r') as f:
    txt = f.read()

classes = ast.literal_eval(txt)
val.names=classes

# id_others = np.where(~ np.isin(np.array(_val.targets), c_to_del))[0]
# others = Subset(_train, id_others)
others = torch.load(os.path.join('dataset_utils','files','others.pt'))

# loss = nn.CrossEntropyLoss(reduction='none')
# optimizer=torch.optim.SGD(model.parameters(), lr=hyperparams['lr'])

def up_trim_mean(a:torch.Tensor):
    return (a.sum() - a.max()) / (len(a) - 1)

def down_trim_mean(a:torch.Tensor):
    return (a.sum() - a.min()) / (len(a) - 1)



#TRAINING LOOP
def test_loop(
        n_epochs,
        model,
        img,
        hyp
):
    global img_idx,immagine,SCORE, CLASSE
    size_others = 1
    others_loader = DataLoader(random_split(others, [size_others, len(others) - size_others])[0], batch_size=64, shuffle=True)

    A,B = model
    A,B=A.cuda(),B.cuda()
    ahook, bhook = {}, {}

    for (an, av), (bn, bv) in zip(A.named_children(), B.named_children()):
        if isinstance(av, nn.Sequential):
            for (an2, av2), (bn2, bv2) in zip(av.named_children(), bv.named_children()):
                if 'basicblock' in an2.lower() or isinstance(av2, models.resnet.BasicBlock):
                    for (an3, av3), (bn3, bv3) in zip(av2.named_children(), bv2.named_children()):
                        ahook[f'{an}_{an2}_{an3}'] = av2.register_forward_hook(get_layer_act(True, f'{an}_{an2}_{an3}'))
                        bhook[f'{bn}_{bn2}_{bn3}'] = bv2.register_forward_hook(get_layer_act(False, f'{bn}_{bn2}_{bn3}'))
                else:
                    ahook[f'{an}_{an2}'] = av2.register_forward_hook(get_layer_act(True, f'{an}_{an2}'))
                    bhook[f'{bn}_{bn2}'] = bv2.register_forward_hook(get_layer_act(False, f'{bn}_{bn2}'))
        else:
            ahook[an] = av.register_forward_hook(get_layer_act(True, an))
            bhook[bn] = bv.register_forward_hook(get_layer_act(False, bn))


    for epoch in range(n_epochs):

        # val_loader = DataLoader(random_split(val, [size_others, len(val) - size_others])[0], batch_size=64, shuffle=True)
        #
        #
        # img_idx = val_loader.dataset.indices
        print(f'Val {epoch=} ')
        A.eval()
        B.eval()

        with torch.no_grad():
            for idx, imgs in enumerate(img):
                immagine = DeT(imgs)
                imgs = imgs.cuda()
                # labels = labels.cuda()

                ascores = A(imgs)
                bscores = B(imgs)
                SCORE=ascores
                # CLASSE = val_loader.dataset.dataset.names[int(bscores.argmax())]
                aent = torch.Tensor([entropy(nn.functional.softmax(ascores, dim=1).cpu().numpy()[i]) for i in range(len(ascores))])
                bent = torch.Tensor([entropy(nn.functional.softmax(bscores, dim=1).cpu().numpy()[i]) for i in range(len(bscores))])
                amean = aent.mean()
                bmean = bent.mean()

                atmean = up_trim_mean(aent)
                btmean = down_trim_mean(bent)

                print(ascores.topk(10).indices)
                print(bscores.topk(10).indices)

                print('Entropy A:', atmean, aent.min(), aent.max(), aent.std())
                print('Entropy B:', btmean, bent.min(), bent.max(), bent.std())
                x=0

                # cl_acc = (t_out_val.max(1).indices == c_to_del).sum() / t_out_val.max(1).indices.shape[0]

            #     ot_t_acc = []
            # for o, ol in others_loader:
            #     o = o.cuda()
            #     ol = ol.cuda()
            #     test_outs = model(o)
            #     ot_t_acc.append((test_outs.max(1).indices == ol).sum() / (test_outs.max(1).indices == ol).shape[0])


        # if epoch == 1 or epoch % 10 == 0:
        #     print(f"Epoch {epoch}, Training loss {loss_train.item():.4f}, Validation loss {loss_val.item():.4f}")

test_loop(
    n_epochs = 1,
    model = [general, unlearnt],
    img=[img],
    hyp=hyperparams
)

def conv_diff(a,b):
    ret = []
    for i in range(a.shape[0]):
        ret.append([
            [a[i].min(),a[i].max(), a[i].mean(), a[i].var()],
            [b[i].min(),b[i].max(), b[i].mean(), b[i].var()]
        ])

        # kl1 = kl_div(a[i], b[i])
        # kl2 = kl_div(b[i], a[i])
        #
        # ret.append([
        #     [kl1.min(), kl1.max(), kl1.mean(), kl1.var()],
        #     [kl2.min(), kl2.max(), kl2.mean(), kl2.var()]
        # ])

    return ret

def others_diff(a,b):
    ret = []
    ret.append([
        [a.min(),a.max(), a.mean(), a.var()],
        [b.min(),b.max(), b.mean(), b.var()]
    ])

    # kl1 = kl_div(a, b)
    # kl2 = kl_div(b, a)
    #
    # ret.append([
    #     [kl1.min(), kl1.max(), kl1.mean(), kl1.var()],
    #     [kl2.min(), kl2.max(), kl2.mean(), kl2.var()]
    # ])

    return ret

def f(a, b, start_row=1,method='diff', N=True):
    global ws

    ret = dict()
    row = start_row
    for i, ((ak, av), (bk, bv)) in enumerate(zip(a.items(), b.items())):
        # print(aa[0], '\t', (aa[1]-bb[1]).sum())
        xa, xb = av.detach().data, bv.detach().data
        xa = xa.cpu() #torch.nn.functional.relu(xa)
        if '28' in ak or 'layer4_1_conv1' in ak:
            import torch.nn.functional as FF
            import matplotlib.pyplot as plt

            plt.imshow(immagine[0].detach().cpu().permute(1, 2, 0).numpy())
            plt.show()

            upsampled_b = FF.interpolate(xb, (224, 224), mode='bilinear', align_corners=False)
            x=0
        xb = xb.cpu() #torch.nn.functional.relu(xb)

        alpha= unlearnt.layer4[1].conv1.alpha # unlearnt.features[28].alpha
        alpha = torch.sigmoid(alpha)

        if False:
            for i in (alpha).topk(10).indices:
                test = (upsampled_b[:, i])[(None,) + (...,)] * _T(immagine).cuda()
                test -= test.min()
                test /= test.max()
                plt.imshow(test[0].permute(1, 2, 0).detach().cpu().numpy(), cmap='jet')
                plt.show()

                test = (upsampled_b[:, i])
                test -= test.min()
                test /= (test.max() + 1e-12)
                plt.imshow(test[0].detach().cpu().numpy(), cmap='jet')
                plt.show()

                r, c = coords
                num = torch.sum(test[int(r / R * img.shape[-1]): int(r / R * img.shape[-1] + img.shape[-2] / R),
                                int(c * (img.shape[-2]) // C): int(c * (img.shape[-2]) // C + img.shape[-1] / C)])
                den = torch.sum(test)
                print(i, round(float(num),2), round(float(den),2), round(float(num / den),2))

            for i in (1-alpha).topk(10).indices:
                test = (upsampled_b[:, i])[(None,) + (...,)] * _T(immagine).cuda()
                test -= test.min()
                test /= test.max()
                plt.imshow(test[0].permute(1, 2, 0).detach().cpu().numpy(), cmap='jet')
                plt.show()

                test = (upsampled_b[:, i]).cuda()
                test -= test.min()
                test /= test.max()
                plt.imshow(test[0].detach().cpu().numpy(), cmap='jet')
                plt.show()

                r, c = coords
                num = torch.sum(test[int(r / R * img.shape[-1]): int(r / R * img.shape[-1] + img.shape[-2] / R),
                                int(c * (img.shape[-2]) // C): int(c * (img.shape[-2]) // C + img.shape[-1] / C)])
                den = torch.sum(test)
                print(i, round(float(num),2), round(float(den),2), round(float(num / den),2))

        # if '29' in ak or True:
        #     try:
        #         path='/homes/spoppi/pycharm_projects/inspecting_twin_models/activations_png/vgg16'
        #         path = os.path.join(path, f'class_{c_to_del}',f'imgidx_{img_idx}', ak)
        #         if not os.path.isdir(path):
        #             os.makedirs(path)
        #         plot_pairs(
        #             xa, xb, 0, ak, idx2=-2,
        #             path=path
        #         )
        #     except Exception as e:
        #         print(e)


        # if 'conv' in ak:
        #     # xa, xb = norm(xa, xb)
        #     val = conv_diff(xa[0], xb[0])
        # else:
        #     val = others_diff(xa, xb)

        # ws[f'A{row}'] = ak
        # row += 1
        # first_row = row
        # ws, row = pyxlutils.write_layer(
        #     val=val,
        #     ws=ws,
        #     st_row=first_row
        #
        # )



f(AH, BH)

ws = pyxlutils.color_scale(ws=ws, cols=['M','N','O','P'])
ws = pyxlutils.adapt_cols_len(ws=ws)
wb.save(out_name)
