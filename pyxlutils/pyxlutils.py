from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import colors

from collections.abc import Iterable

def create_ws(
        ws,
        cname
):
    ws['A1'] = cname
    ws.merge_cells(range_string='C1:F1')
    ws['C1'] = 'Standard'
    ws['C1'].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(range_string='H1:K1')
    ws['H1'] = 'Modified'
    ws['H1'].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(range_string='M1:P1')
    ws['M1'].alignment = Alignment(horizontal="center", vertical="center")
    ws['M1'] = 'Diff'

    ws['C2'] = 'min'
    ws['D2'] = 'max'
    ws['E2'] = 'mean'
    ws['F2'] = 'std'
    ws['H2'] = 'min'
    ws['I2'] = 'max'
    ws['J2'] = 'mean'
    ws['K2'] = 'std'
    ws['M2'] = 'min'
    ws['N2'] = 'max'
    ws['O2'] = 'mean'
    ws['P2'] = 'std'

    return ws

def create_ws_2(
        ws,
        cname
):
    ws['A1'] = cname
    ws.merge_cells(range_string='C1:F1')
    ws['C1'] = 'Unlearnt'
    ws['C1'].alignment = Alignment(horizontal="center", vertical="center")

    return ws

def write_layer(
        ws,
        val,
        st_row
):
    row = st_row
    for idx, v in enumerate(val):
        col = 2
        ws[f'{get_column_letter(col)}{row}'] = idx
        col += 1
        v.append([abs(a - b) for a, b in zip(*v)])
        for vv in v:
            for t in vv:
                ws[f'{get_column_letter(col)}{row}'] = float(t.detach().cpu().data.numpy())
                col += 1

            col += 1
        row += 1

    return ws, row

def color_scale(
        ws,
        cols = [],
        rule = None
):
    if rule == None:
        rule = ColorScaleRule(
            start_type='min', start_color='FFFF0000',
            end_type='max', end_color='00FFFFFF'
        )

    assert isinstance(cols, Iterable), f'cols must be iterable.'

    for col in cols:
        ws.conditional_formatting.add(f"{col}4:{col}{len(ws[col])}", rule)

    return ws


def as_text(value):
    if value is None:
        return ""
    return str(value)

def adapt_cols_len(ws):
    for column_cells in ws.columns:
        length = max(len(as_text(cell.value)) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length

    return ws

def write_alpha_layer(
        ws,
        val,
        st_col
):
    col = st_col
    for idx, v in enumerate(val):
        row = 4
        for vv in v:
            for t in vv:
                for tt in t.detach().cpu().numpy():
                    ws[f'{get_column_letter(col)}{row}'] = float(tt)
                    row += 1

    return ws, col+1
