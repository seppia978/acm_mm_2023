import os
import numpy as np
import torch
import ast
import torch.nn as nn
import torch.utils.data as data
from torchvision.utils import make_grid
import torchvision.models as models
import torch.nn.functional as FF
import torchvision.transforms as transforms
from torchvision.datasets import ImageFolder
from torch.utils.data import DataLoader, random_split, Subset, WeightedRandomSampler
from pl_bolts.transforms.dataset_normalizations import cifar10_normalization
from pl_bolts.datamodules import CIFAR10DataModule
from torchmetrics.functional import accuracy
from generics.plotutils import plot_pairs
import pyxlutils
from scipy.special import kl_div
from scipy.stats import entropy
import matplotlib.pyplot as plt


from sklearn.feature_selection import mutual_info_classif as MI

import torchvision.transforms.functional as F
from torchvision.datasets import ImageFolder,\
    CIFAR10, \
    CIFAR100, \
    MNIST
import pandas as pd
import openpyxl as xls
from openpyxl import Workbook
import custom_archs.WTFCNN as WTFCNN

from custom_archs import convert_conv2d_to_alpha, get_all_alpha_layers, set_label,\
    invert_alphas, normalize_alphas
from openpyxl.utils import get_column_letter
import CKA.cka as cka

from unl_shorter_dataset_2 import generate_random_idxs

c_to_del = [3]
immagine, SCORE, CLASSE, LABELS = None, 0, '', []
# which_net = [159, 168, 180, 242, 163, 243, 246, 211, 179, 236]
which_net = [2, 3, 4, 5, 394, 6, 395, 391, 389, 149]
img_idx = -1
AH={}
BH={}
def get_layer_act(is_a, name):
    def register_hook(module, inp, out):
        if is_a:
            AH[name] = out.data
        else:
            BH[name] = out.data
    return register_hook




# PATH = f"/homes/spoppi/pycharm_projects/inspecting_twin_models/checkpoints/short/class_{'-'.join(str(c) for c in c_to_del)}_freeze-before-28_untraining_vgg16.pt"
# PATH = f"/homes/spoppi/pycharm_projects/inspecting_twin_models/checkpoints/short/class_100-classes_freeze-0_untraining_vgg16.pt"
# root = '/work/dnai_explainability/unlearning/icml2023/alpha_matrices/alpha_resnet18-100-0_resnet18_1.0_100.0_2022-12-13-30'
# # PATH = f"{root}/one_for_all_untraining-loss_type_3way_multiplication-ur_10.0-model_resnet18-lambda1_0.1-lambda2_0.1-g_0.01-alpha_init_5.0.pt"
# # PATH = f"{root}/one_for_all_untraining-loss_type_3way_sum-ur_100.0-model_resnet18-lambda1_10.0-lambda2_1.0-g_0.01-alpha_init_3.0-batch_size_64-zero_grad_frequency_8.pt"
# # PATH = f"{root}/one_for_all_untraining-loss_type_3way_sum-ur_100.0-model_resnet18-lambda1_10.0-lambda2_1.0-g_0.01-alpha_init_3.0-batch_size_64-zero_grad_frequency_16-max_val_size_20000.pt"
# PATH = f"{root}/final.pt"
# root = '/work/dnai_explainability/unlearning/icml2023/alpha_matrices/alpha_resnet18-100-0_resnet18_1.0_100.0_2022-12-13-30'



# # PATH = f"{root}/resnet18_1.0_100.0_class-{c_to_del[0]}_alpha.pt"
# net='vgg16' if 'vgg16' in PATH else 'resnet18'
# # net='resnet18'

dataset = 'imagenet'

print(f'cuda: {torch.cuda.is_available()}')

print('Preparing dataset...')

if dataset.lower() == 'imagenet':

    size = 224
    means = [0.485, 0.456, 0.406]
    stds = [0.229, 0.224, 0.225]
    T = transforms.Compose([
        transforms.Resize(256),
        transforms.CenterCrop(size),
        transforms.ToTensor(),
        transforms.Normalize(means, stds)
    ])

    _T = transforms.Compose([
        transforms.Resize(256),
        transforms.CenterCrop(size),
        transforms.Normalize(means, stds)
    ])

    DeT = transforms.Compose([
        transforms.Normalize(-1 * torch.Tensor(means) / torch.Tensor(stds), 1.0 / torch.Tensor(stds))
    ])

    # _train = ImageFolder(
    #     root='/nas/softechict-nas-2/datasets/Imagenet_new/ILSVRC/Data/CLS-LOC/train',
    #     transform=T
    # )

    _val = ImageFolder(
        root='/nas/softechict-nas-2/datasets/Imagenet_new/ILSVRC/Data/CLS-LOC/val',
        transform=T
    )

    root='/mnt/beegfs/work/dnai_explainability/unlearning/icml2023/alpha_matrices/alpha_resnet18-100-0_resnet18_1.0_100.0_2022-12-13-30/alpha_matrices/alpha_resnet18-100-0_resnet18_1.0_100.0_2022-12-19-16'# root = '/mnt/beegfs/work/dnai_explainability/unlearning/icml2023/alpha_matrices/alpha_resnet18-100-0_resnet18_1.0_100.0_2022-12-13-30/alpha_matrices/alpha_resnet18-100-0_resnet18_1.0_100.0_2022-12-19-16/alpha_matrices/test_all_resnet18_1.0_100.0_2022-12-23-22'
    general = models.resnet18(weights='ResNet18_Weights.DEFAULT')

elif dataset.lower() == 'cifar10':

    # means = [0.5, 0.5, 0.5]
    # stds = [0.5, 0.5, 0.5]

    mean=[x / 255.0 for x in [125.3, 123.0, 113.9]]
    std=[x / 255.0 for x in [63.0, 62.1, 66.7]]
    T = transforms.Compose(
        [
            transforms.ToTensor(),
            cifar10_normalization(),
        ]
    )

    DeT = transforms.Compose(
        [
            transforms.Normalize(
                -1. * torch.Tensor(mean) / torch.Tensor(std), 1. / torch.Tensor(std)
            )
        ]
    )

    _train = CIFAR10(
        root='/work/dnai_explainability/unlearning/datasets/cifar10_classification/data',
        transform=T, download=True, train=True
    )

    _val = CIFAR10(
        root='/work/dnai_explainability/unlearning/datasets/cifar10_classification/data',
        transform=T, download=True, train=False
    )

    root = '/mnt/beegfs/work/dnai_explainability/unlearning/icml2023/alpha_matrices/alpha_resnet18-100-0_resnet18_1.0_100.0_2022-12-13-30/alpha_matrices/alpha_resnet18-100-0_resnet18_1.0_100.0_2022-12-19-16/alpha_matrices/alpha_resnet18-100-0_resnet18_1.0_100.0_2022-12-29-48'

    general = models.resnet18(pretrained=False, num_classes=10)
    # general.conv1 = nn.Conv2d(3, 64, kernel_size=(7, 7), stride=(1, 1), padding=(1, 1), bias=False)
    acab = torch.load(
        os.path.join(
            '/work/dnai_explainability/unlearning/datasets/cifar10_classification/checkpoints',
            '2022-12-29_5.pt'
        )
    )
    ac=list(map(lambda x: x[6:], acab.keys()))
    ckp = dict()
    for k1,k2 in zip(acab,ac):
        if k2 == k1[6:]:
            ckp[k2] = acab[k1]
    general.load_state_dict(ckp)

elif dataset.lower() == 'cifar100':

    # T = transforms.Compose(
    #     [
    #         transforms.ToTensor(),
    #         cifar10_normalization(),
    #     ]
    # )

    means = [0.5, 0.5, 0.5]
    stds = [0.5, 0.5, 0.5]
    T = transforms.Compose([
        transforms.ToTensor(),
        transforms.Normalize(means, stds)
    ])

    # _train = CIFAR100(
    #     root='/work/dnai_explainability/unlearning/datasets/cifar100_classification/train',
    #     transform=T, download=True, train=True
    # )

    _val = CIFAR100(
        root='/work/dnai_explainability/unlearning/datasets/cifar100_classification/val',
        transform=T, download=True, train=False
    )

elif dataset.lower() == 'mnist':

    means, stds = (0.1307,), (0.3081,)
    T  = transforms.Compose([
        transforms.ToTensor(),
        transforms.Normalize(means, stds)
    ])

    # _train = MNIST(
    #     root='/work/dnai_explainability/unlearning/datasets/mnist_classification/train',
    #     transform=T, download=True, train=True
    # )

    _val = MNIST(
        root='/work/dnai_explainability/unlearning/datasets/mnist_classification/val',
        transform=T, download=True, train=False
    )

    print('Creating models...')

    root = '//mnt/beegfs/work/dnai_explainability/unlearning/icml2023/alpha_matrices/alpha_resnet18-100-0_resnet18_1.0_100.0_2022-12-13-30/alpha_matrices/alpha_resnet18-100-0_resnet18_1.0_100.0_2022-12-19-16/alpha_matrices/test_all_resnet18_1.0_100.0_2022-12-28-24'

    general = models.resnet18(pretrained=False, num_classes=10)
    general.conv1 = nn.Conv2d(1, 64, kernel_size=(7, 7), stride=(2, 2), padding=(3, 3), bias=False)
    acab = torch.load(
        os.path.join(
            '/work/dnai_explainability/unlearning/datasets/mnist_classification/checkpoints',
            '2022-12-28_0.pt'
        )
    )
    # acab=acab['state_dict']
    ac=list(map(lambda x: x[6:], acab.keys()))
    ckp = dict()
    for k1,k2 in zip(acab,ac):
        if k2 == k1[6:]:
            ckp[k2] = acab[k1]
    general.load_state_dict(ckp)


classes_number = len(_val.classes)

PATH = f"{root}/final.pt"
net='resnet18'

if net == 'resnet18':
    # general.conv1 = nn.Conv2d(1, 64, kernel_size=(7, 7), stride=(2, 2), padding=(3, 3), bias=False)
    # general = convert_conv2d_to_alpha(models.resnet18(pretrained=False), m=10, standard=False)
    # unlearnt = convert_conv2d_to_alpha(models.resnet18(pretrained=False))
    unlearnt = WTFCNN.WTFCNN(
        kind=WTFCNN.resnet18, pretrained=True,
        m=3., classes_number=classes_number, resume=PATH,
        dataset=dataset.lower()
    )
    # C = convert_conv2d_to_alpha(models.1resnet18(pretrained=False))
elif net == 'vgg16':
    general = convert_conv2d_to_alpha(models.vgg16(pretrained=True), m=10, standard=False)
    unlearnt = convert_conv2d_to_alpha(models.vgg16(pretrained=False))



# unlearnt.load_state_dict(torch.load(PATH))
# C.load_state_dict(torch.load(PATH))
# C = invert_alphas(C)
#E = normalize_alphas(E)
# PATH = f"/homes/spoppi/pycharm_projects/inspecting_twin_models/checkpoints/all_classes/one_for_all_untraining.pt"
# general.load_state_dict(torch.load(PATH))
# au, ag = get_all_alpha_layers(unlearnt), get_all_alpha_layers(general)
# matrix = cka.CKA(au, ag)
# general.load_state_dict(torch.load(PATH))
# invert_alphas(unlearnt)

# wb = Workbook()
#
# ws = wb.active
# ws = pyxlutils.create_ws(ws=ws, cname='shark')




out_name = f'activations_{"-".join(str(c) for c in c_to_del)}_vgg16.xlsx'

hyperparams = {
    'lr': 1e-3,
    'model': 'resnet18'
}

device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')

# size=224
# means = [0.485, 0.456, 0.406]
# stds = [0.229, 0.224, 0.225]

# T = transforms.Compose([
#     transforms.Resize(256),
#     transforms.CenterCrop(size),
#     transforms.ToTensor(),
#     transforms.Normalize(means, stds)
# ])

# _T = transforms.Compose([
#     transforms.Resize(256),
#     transforms.CenterCrop(size),
#     transforms.Normalize(means, stds)
# ])

# DeT = transforms.Compose([
#     transforms.Normalize(-1 * torch.Tensor(means) / torch.Tensor(stds), 1.0 / torch.Tensor(stds))
# ])


# _train = ImageFolder(
#     root='/nas/softechict-nas-2/datasets/Imagenet_new/ILSVRC/Data/CLS-LOC/train',
#     transform=T
# )
#
# id_to_remove = np.where(np.isin(np.array(_train.targets), c_to_del))[0]
#
#
# train = Subset(_train, id_to_remove)
# train = torch.load(os.path.join('dataset_utils','files','train.pt'))


# _val = ImageFolder(
#     root='/nas/softechict-nas-2/datasets/Imagenet_new/ILSVRC/Data/CLS-LOC/val',
#     transform=T
# )
#
# id_to_remove = np.where(np.isin(np.array(_val.targets), c_to_del))[0]
#
# val = Subset(_val, id_to_remove)
# _val = torch.load(os.path.join('dataset_utils','files','val.pt'))
id_c = np.where(np.isin(np.array(_val.targets), c_to_del))[0]
id_others = np.where(~ np.isin(np.array(_val.targets), c_to_del))[0]

val_c = Subset(_val, id_c)
val_c.targets = torch.Tensor(_val.targets).int()[id_c].tolist()
val_others = Subset(_val, id_others)
val_others.targets = torch.Tensor(_val.targets).int()[id_others].tolist()

concat_train = data.ConcatDataset((val_c,val_others))
concat_train.targets = torch.Tensor(_val.targets).int()[id_c].tolist()
concat_train.targets = [*val_c.targets, *val_others.targets]

# perm_c = torch.randperm(torch.Tensor(id_c).shape[0])
# idx_c = perm_c[0]
# idx = id_c[idx_c]
#
# perm_ot = torch.randperm(torch.Tensor(id_others).shape[0])
# idx_ot = perm_ot[:3]
# idx.extend(id_others[idx_ot])

# val = Subset(_val.dataset, id_c)

# val_c = Subset(_val.dataset, id_c)
# val_ot = Subset(_val.dataset, id_others)
# val = (val_c,val_ot)

# test = ImageFolder(
#     root='/nas/softechict-nas-2/datasets/Imagenet_new/ILSVRC/Data/CLS-LOC/test',
#     transform=T
# )

#
with open('class_names/names.txt',  'r') as f:
    txt = f.read()

classes = ast.literal_eval(txt)
concat_train.names=classes

lab=c_to_del[0]
a = unlearnt.arch.layer1[0].conv1.alpha[lab]
_a=1-torch.sigmoid(a)
a_=torch.sigmoid(a)
minidx = torch.argwhere(
    unlearnt.arch.layer1[0].conv1.alpha[lab]<-2
)


# id_others = np.where(~ np.isin(np.array(_val.targets), c_to_del))[0]
# others = Subset(_train, id_others)
# others = torch.load(os.path.join('dataset_utils','files','others.pt'))

# loss = nn.CrossEntropyLoss(reduction='none')
# optimizer=torch.optim.SGD(model.parameters(), lr=hyperparams['lr'])

def up_trim_mean(a:torch.Tensor):
    return (a.sum() - a.max()) / (len(a) - 1)

def down_trim_mean(a:torch.Tensor):
    return (a.sum() - a.min()) / (len(a) - 1)


print('Testing...')
#TRAINING LOOP
def test_loop(
        n_epochs,
        optimizer,
        model,
        loss_fn,
        val,
        hyp
):
    global img_idx,immagine,SCORE, CLASSE, LABELS
    size_others = 10
    # others_loader = DataLoader(random_split(others, [size_others, len(others) - size_others])[0], batch_size=64, shuffle=True)

    A,B = model
    t1 = WTFCNN.WTFCNN(
        kind=WTFCNN.resnet18, pretrained=True,
        m=3., classes_number=classes_number, resume=PATH,
        dataset=dataset.lower()
    )
    t2 = WTFCNN.WTFCNN(
        kind=WTFCNN.resnet18, pretrained=True,
        m=3., classes_number=classes_number, resume=None,
        dataset=dataset.lower()
    )
    A.cuda(),B.arch.cuda()
    t1.arch.cuda()
    t2.arch.cuda()
    ahook, bhook = {}, {}

    for (an, av), (bn, bv) in zip(A.named_children(), B.arch.named_children()):
        if isinstance(av, nn.Sequential):
            for (an2, av2), (bn2, bv2) in zip(av.named_children(), bv.named_children()):
                if 'basicblock' in an2.lower() or isinstance(av2, models.resnet.BasicBlock):
                    for (an3, av3), (bn3, bv3) in zip(av2.named_children(), bv2.named_children()):
                        ahook[f'{an}{an2}{an3}'] = av2.register_forward_hook(get_layer_act(True, f'{an}{an2}{an3}'))
                        bhook[f'{bn}{bn2}{bn3}'] = bv2.register_forward_hook(get_layer_act(False, f'{bn}{bn2}{bn3}'))
                else:
                    ahook[f'{an}{an2}'] = av2.register_forward_hook(get_layer_act(True, f'{an}{an2}'))
                    bhook[f'{bn}{bn2}'] = bv2.register_forward_hook(get_layer_act(False, f'{bn}{bn2}'))
        else:
            ahook[an] = av.register_forward_hook(get_layer_act(True, an))
            bhook[bn] = bv.register_forward_hook(get_layer_act(False, bn))

    # y_train = val.targets
    # weight = 1. / torch.Tensor([.001 for _ in range(1000)])
    #
    # weight[~np.isin(list(range(len(weight))), np.array(c_to_del))] = .5 / (len(weight) - len(c_to_del))
    # weight[np.array(c_to_del)] = .5 / len(c_to_del)
    #
    # samples_weight = np.array([weight[t] for t in y_train])
    # samples_weight = torch.from_numpy(samples_weight)
    # sampler = WeightedRandomSampler(samples_weight.type('torch.DoubleTensor'), len(samples_weight))

    c_to_del, c2 = [444], [2]

    id_c = np.where(np.isin(np.array(val.targets), c_to_del))[0]
    id_others = np.where(np.isin(np.array(val.targets), c2))[0]
    # id_others = np.where(~ np.isin(np.array(_val.targets), c_to_del))[0]

    val_c = Subset(val, id_c)
    val_c.targets = torch.Tensor(val.targets).int()[id_c].tolist()
    val_others = Subset(val, id_others)
    val_others.targets = torch.Tensor(val.targets).int()[id_others].tolist()

    concat_train = data.ConcatDataset((val_c,val_others))
    concat_train.targets = torch.Tensor(_val.targets).int()[id_c].tolist()
    concat_train.targets = [*val_c.targets, *val_others.targets]

    y_train = concat_train.targets  # train.datasets[0].dataset.targets

    weight = 1. / torch.Tensor([len(np.unique(y_train))/50 for _ in range(len(concat_train))])

    # weight[np.isin(list(range(len(weight))), np.array(c_to_del))] = .5 / (len(c_to_del))
    # weight[np.array(c2)] = .5 / len(c2)

    samples_weight = weight # np.array([weight[t] for t in y_train])
    # samples_weight = torch.from_numpy(samples_weight)
    sampler = WeightedRandomSampler(samples_weight.type('torch.DoubleTensor'), len(samples_weight))


    val_loader = DataLoader(val, batch_size=size_others, num_workers=4, shuffle=True)
    # val_loader = DataLoader(concat_train, batch_size=size_others, num_workers=4, sampler=sampler)

    for epoch in range(n_epochs):

        # val_loader = DataLoader(Subset(val,[13]), batch_size=64, shuffle=True)

        ##########################################################################


        ##########################################################################

        # val_loader = DataLoader(random_split(val, [size_others, len(val) - size_others])[0], batch_size=64, shuffle=True)

        # img_idx = val_loader.dataset.indices
        print(f'Val {epoch=} ')
        A.eval()
        B.arch.eval()
        t1.arch.eval()
        t2.arch.eval()
        # C.eval()

        with torch.no_grad():
            for idx, (imgs, labels) in enumerate(val_loader):
                # immagine = DeT(imgs)
                imgs = imgs.cuda()
                LABELS = labels.tolist()
                labels = labels.cuda()

                # set_label(B,labels)
                # set_label(C,labels)


                # B.set_label(labels.long())
                # from torchcam.methods import ScoreCAM
                # uCAM=ScoreCAM(B.arch)
                # gCAM=ScoreCAM(A)

                # ascores = A(imgs)
                # bscores = B(imgs, labels=labels.long())

                # uact=uCAM(ascores.argmax().item(), bscores)
                # gact=gCAM(ascores.argmax().item(), ascores)

                # import matplotlib.pyplot as plt
                # from torchcam.utils import overlay_mask
                # img=DeT(imgs[0].cpu().detach())
                # gresult = overlay_mask(F.to_pil_image(img), F.to_pil_image(gact[0].squeeze(0), mode='F'), alpha=0.5)
                # uresult = overlay_mask(F.to_pil_image(img), F.to_pil_image(uact[0].squeeze(0), mode='F'), alpha=0.5)
                # plt.imshow(uresult); plt.axis('off'); plt.tight_layout();
                # plt.savefig('resu.png')
                # plt.imshow(gresult); plt.axis('off'); plt.tight_layout();
                # plt.savefig('resg.png')

                from torchcam.methods.activation import ScoreCAM
                cam=ScoreCAM(A)
                for im in range(size_others):
                    ascores = A(imgs[im].unsqueeze(0))


                    def attr_method(image=None, alpha=None, activations=None, classidx=None):

                                                    
                                        ups = [activations[x] for x in activations.keys() if 'conv' in x]
                                        ups = [(x-x.min())/(x.max()-x.min()+1e-15) for x in ups]

                                        w = list(alpha.get_all_alpha_layers().values())
                                        w = [1-torch.sigmoid(x) for x in w]
                                        import torch.nn.functional as FF
                                        s=[]
                                        for i in range(len(ups)):    
                                            s.append(ups[i]*w[i][0][(None,)+(...,)+(None,)*2])
                                        s0 = [x[0].unsqueeze(0) for x in s]
                                        s0sum = [x.sum(axis=1) for x in s0]
                                        s0sumups = [FF.interpolate(x.unsqueeze(0),[224,224],mode='bilinear') for x in s0sum]
                                        w444 = [x[0] for x in w]
                                        sal = [x*w_.max() for x,w_ in zip(s0sumups,w444)]
                                        sal = torch.cat(sal, dim=1)
                                        sal = sal.sum(axis=(0,1))
                                        sal-=sal.min()
                                        sal/=sal.max()

                                        return sal.squeeze(0).to('cuda')
                                    # plt.imshow((0.3*DeT(imgs[im]) + 0.7 * sal).permute(1,2,0).detach().cpu())
                                    # plt.savefig('sal.png')
                    mnsal=attr_method(
                                        image=imgs[im].unsqueeze(0),
                                        activations=AH,
                                        alpha=B
                                    )
                    img=DeT(imgs[im]).cpu().permute(1,2,0).detach()
                    mnsal=mnsal.cpu().detach()
                    fig = plt.figure()
                    #fig.set_size_inches(width / height, 1, forward=False)
                    ax = plt.Axes(fig, [0., 0., 1., 1.])
                    ax.set_axis_off()
                    fig.add_axes(ax)
                    my_cmap1 = plt.cm.get_cmap('jet')
                    ax.imshow(img, origin='upper')
                    ax.imshow(mnsal, origin='upper', extent=[0, img.shape[1], img.shape[0], 0], alpha=0.5,cmap=my_cmap1)
                    plt.savefig(f'''explanation_map_1.png''')
                    plt.close(fig)
                    print(classes[int(labels[im])])


                    scsal=cam(class_idx=[labels[0]], scores=ascores)
                    scsal=FF.interpolate(scsal[0].unsqueeze(0), (224,224), mode='bilinear').squeeze()
                    scsal-=scsal.min()
                    scsal/=scsal.max()
                    # plt.imshow((0.3*DeT(imgs[im]) + 0.7 * sal).permute(1,2,0).detach().cpu())
                    # plt.savefig('sal.png')

                    img=DeT(imgs[im]).cpu().permute(1,2,0).detach()
                    scsal=scsal.cpu().detach()
                    fig = plt.figure()
                    #fig.set_size_inches(width / height, 1, forward=False)
                    ax = plt.Axes(fig, [0., 0., 1., 1.])
                    ax.set_axis_off()
                    fig.add_axes(ax)
                    my_cmap1 = plt.cm.get_cmap('jet')
                    ax.imshow(img, origin='upper')
                    ax.imshow(scsal, origin='upper', extent=[0, img.shape[1], img.shape[0], 0], alpha=0.5,cmap=my_cmap1)
                    plt.savefig(f'''scam_explanation_map_1.png''')
                    plt.close(fig)
                    print(classes[int(labels[im])])
                    # scsal=cam(class_idx=[labels[0]], scores=ascores)
                    # scsal=FF.interpolate(scsal[0].unsqueeze(0), (224,224), mode='bilinear').squeeze()
                    # scsal-=scsal.min()
                    # scsal/=scsal.max()

                    from ADCC.metrics import average_drop, coherency, complexity
                    mineavgdrop=average_drop.average_drop(
                                        image=imgs[im].unsqueeze(0), 
                                        explanation_map=(
                                            imgs[im].unsqueeze(0)*mnsal[(None,)*2+(...,)].cuda()
                                        ),
                                        arch=A, 
                                        out=torch.softmax(ascores,1), 
                                        class_idx=labels[im].unsqueeze(0)
                    )
                    scamavgdrop=average_drop.average_drop(
                                        image=imgs[im].unsqueeze(0), 
                                        explanation_map=(
                                            imgs[im].unsqueeze(0)*scsal[(None,)*2+(...,)].cuda()
                                        ),
                                        arch=A, 
                                        out=torch.softmax(ascores,1), 
                                        class_idx=labels[im].unsqueeze(0)
                    )

                    
                    minecom=mnsal.mean()
                    scamcom=scsal.mean()
                    # print(mnsal.mean(), scsal.mean())

                    from scipy.stats import pearsonr as pcc
                    def coherency(saliency_map, sal1, arch, attr_method, out):
                                        if torch.cuda.is_available():
                                            sal1 = sal1.cuda()

                                        A, B = saliency_map.detach(), sal1.detach()

                                        '''
                                        # Pearson correlation coefficient
                                        # '''
                                        Asq, Bsq = A.view(1, -1).squeeze(0).cpu(), B.view(1, -1).squeeze(0).cpu()

                                        import os
                                        os.environ['KMP_DUPLICATE_LIB_OK'] = 'True'

                                        if torch.tensor(Asq).isnan().any() or torch.tensor(Bsq).isnan().any():
                                            y = 0.
                                        else:
                                            y, _ = pcc(Asq, Bsq)
                                            y = (y + 1) / 2




                                        return y,A,B


                    A(imgs[im].unsqueeze(0)*mnsal[(None,)*2+(...,)].cuda())
                    mnsal1= attr_method(
                                        image=imgs[im].unsqueeze(0)*mnsal[(None,)*2+(...,)].cuda(),
                                        activations=AH,
                                        alpha=B
                    )

                    A(imgs[im].unsqueeze(0)*scsal[(None,)*2+(...,)].cuda())
                    scsal1=cam(class_idx=[labels[0]], scores=ascores)
                    scsal1=FF.interpolate(scsal1[0].unsqueeze(0), (224,224), mode='bilinear').squeeze()
                    scsal1-=scsal1.min()
                    scsal1/=scsal1.max()

                    minecoh=coherency(mnsal, mnsal1,B,attr_method,ascores)[0]
                    scamcoh=coherency(scsal, scsal1,B,attr_method,ascores)[0]

                    adccmn = 3 / (1/minecoh + 1/(1-minecom) +1/(1-mineavgdrop))
                    adccsc = 3 / (1/scamcoh + 1/(1-scamcom) +1/(1-scamavgdrop))
                    
                    print(mineavgdrop, scamavgdrop)
                    print(minecom, scamcom)
                    print(minecoh, scamcoh)
                    print(adccmn, adccsc)


                t1scores = t1(imgs, labels=labels)
                # cscores = C(imgs)

                # CLASS = torch.Tensor([
                #     (torch.rand(1) * classes_number) \
                #     for _ in range(labels.shape[0])
                # ]).long()
                # set_label(B, CLASS.cuda())

                # kCAM=ScoreCAM(B.arch)

                labs_portion = labels.clone()
                CLASS = generate_random_idxs(
                    labs_portion, 1, classes_number
                ).squeeze()

                dscores = B(imgs, labels=CLASS.cuda())
                # kact=kCAM(dscores.argmax().item(), dscores)

                # kresult = overlay_mask(F.to_pil_image(img), F.to_pil_image(kact[0].squeeze(0), mode='F'), alpha=0.5)
                # plt.imshow(kresult); plt.axis('off'); plt.tight_layout();
                # plt.savefig('resk.png')

                # act=(AH['layer4_1_conv2'] * (1-torch.sigmoid(B.arch.layer4[1].conv2.alpha[labels])[(...,) + (None,) *2])).sum(dim=1)
                # act-=act.min()
                # act/=(act.max()+1e-8)
                # hresult = overlay_mask(F.to_pil_image(img), F.to_pil_image(act[0].squeeze(0), mode='F'), alpha=0.5)
                # plt.imshow(hresult); plt.axis('off'); plt.tight_layout();
                # plt.savefig('resh.png')

                CLASS = (labels+1) % classes_number
                escores = B(imgs, labels=CLASS.cuda())

                print(labels)
                # print(f"Predictions pretrained net: {ascores.max(1).indices}")
                # print(f"Predictions unlearnt net: {bscores.max(1).indices}")
                # # print(f"Predictions inverted net: {cscores.max(1).indices}")
                # print(f"Predictions kept net: {dscores.max(1).indices}")
                # print(f"Predictions kept2 net: {escores.max(1).indices}")

                # print(
                #     f'Score pretrained net: {accuracy(ascores, labels, task="multiclass", num_classes=classes_number)}\n',
                #     f'Score unlearnt net: {accuracy(bscores, labels, task="multiclass", num_classes=classes_number)}\n',
                #     # f'Score inverted net: {round(float((labels == cscores.max(1).indices).sum()/labels.shape[0]),3)}\n',
                #     f'Score kept net: {accuracy(dscores, labels, task="multiclass", num_classes=classes_number)}\n',
                #     f'Score kept2 net: {accuracy(escores, labels, task="multiclass", num_classes=classes_number)}\n'
                # )

                idx = torch.argwhere(B.arch.layer3[0].conv1.alpha[444] == -3)



                base = torch.softmax(B(imgs[0].unsqueeze(0), torch.Tensor([444]).long()),1)[:,444]
                scarti_0, avg0 = {}, 0.
                scarti_1, avg1 = {}, 1.
                print(labels[0])
                it=10

                i3 = torch.argwhere(labels == 444)
                for iii in i3:
                    t1 = WTFCNN.WTFCNN(
                        kind=WTFCNN.resnet18, pretrained=True,
                        m=3., classes_number=classes_number, resume=PATH,
                        dataset=dataset.lower()
                    )
                    t1.arch.cuda().eval()

                    for ii in range(it):
                        if ii in scarti_0.keys():
                            scarti_0[ii] += (
                                float(torch.softmax(
                                    t1(imgs[iii], torch.Tensor([444]).long()),
                                    1
                                )[:,444])
                            )
                        else:
                            scarti_0[ii] = (
                                float(torch.softmax(
                                    t1(imgs[iii], torch.Tensor([444]).long()),
                                    1
                                )[:,444])
                            )
                        if ii in scarti_1.keys():
                            scarti_1[ii] += (
                                float(torch.softmax(
                                    t1(imgs[torch.argwhere(labels == 444)[0]], torch.Tensor([444]).long()),
                                    1
                                )[:,2])
                            )
                        else:
                            scarti_1[ii] = (
                                float(torch.softmax(
                                    t1(imgs[torch.argwhere(labels == 444)[0]], torch.Tensor([444]).long()),
                                    1
                                )[:,2])
                            )
                        t1.set_n_alphas(444, n=1/it)

                # print(scarti)


                SCORE=ascores
                # break
                # # CLASSE = val_loader.dataset.dataset.names[int(bscores.argmax())]
                # aent = torch.Tensor([entropy(nn.functional.softmax(ascores, dim=1).cpu().numpy()[i]) for i in range(len(ascores))])
                # bent = torch.Tensor([entropy(nn.functional.softmax(bscores, dim=1).cpu().numpy()[i]) for i in range(len(bscores))])
                # amean = aent.mean()
                # bmean = bent.mean()
                #
                # atmean = up_trim_mean(aent)
                # btmean = down_trim_mean(bent)

                # print(ascores.topk(10).indices)
                # print(bscores.topk(10).indices)
                #
                # print('Entropy A:', atmean, aent.min(), aent.max(), aent.std())
                # print('Entropy B:', btmean, bent.min(), bent.max(), bent.std())
                # break
                # x=0

                # cl_acc = (t_out_val.max(1).indices == c_to_del).sum() / t_out_val.max(1).indices.shape[0]

            #     ot_t_acc = []
            # for o, ol in others_loader:
            #     o = o.cuda()
            #     ol = ol.cuda()
            #     test_outs = model(o)
            #     ot_t_acc.append((test_outs.max(1).indices == ol).sum() / (test_outs.max(1).indices == ol).shape[0])


        # if epoch == 1 or epoch % 10 == 0:
        #     print(f"Epoch {epoch}, Training loss {loss_train.item():.4f}, Validation loss {loss_val.item():.4f}")

test_loop(
    n_epochs = 1,
    optimizer = None,
    model = [general, unlearnt],
    loss_fn = None,
    val=_val,
    hyp=hyperparams
)

def conv_diff(a,b):
    ret = []
    for i in range(a.shape[0]):
        ret.append([
            [a[i].min(),a[i].max(), a[i].mean(), a[i].var()],
            [b[i].min(),b[i].max(), b[i].mean(), b[i].var()]
        ])

        # kl1 = kl_div(a[i], b[i])
        # kl2 = kl_div(b[i], a[i])
        #
        # ret.append([
        #     [kl1.min(), kl1.max(), kl1.mean(), kl1.var()],
        #     [kl2.min(), kl2.max(), kl2.mean(), kl2.var()]
        # ])

    return ret

def others_diff(a,b):
    ret = []
    ret.append([
        [a.min(),a.max(), a.mean(), a.var()],
        [b.min(),b.max(), b.mean(), b.var()]
    ])

    # kl1 = kl_div(a, b)
    # kl2 = kl_div(b, a)
    #
    # ret.append([
    #     [kl1.min(), kl1.max(), kl1.mean(), kl1.var()],
    #     [kl2.min(), kl2.max(), kl2.mean(), kl2.var()]
    # ])

    return ret

def f(a, b, start_row=1,method='diff', N=True):
    global ws

    ret = dict()
    row = start_row
    for i, ((ak, av), (bk, bv)) in enumerate(zip(a.items(), b.items())):
        # print(aa[0], '\t', (aa[1]-bb[1]).sum())
        xa, xb = av.detach().data, bv.detach().data
        xa = xa.cpu() #torch.nn.functional.relu(xa)
        if '28' in ak or 'conv1' in ak:
            import torch.nn.functional as FF
            import matplotlib.pyplot as plt
            from CKA import cka

            # cka.distance(torch.Tensor([[1,2],[3,4]]).numpy(),torch.Tensor([[4,3],[2,1]]).numpy())
            # cka.distance(torch.sigmoid(alpha).cpu().detach().numpy(), torch.sigmoid(alpha).cpu().detach().numpy())
            # cka.distance(torch.sigmoid(alpha).cpu().detach().numpy(), torch.sigmoid(1-alpha).cpu().detach().numpy())

            plt.imshow(immagine[0].detach().cpu().permute(1, 2, 0).numpy())
            plt.show()

            upsampled_b = FF.interpolate(xb, (224, 224), mode='bilinear', align_corners=False)
            upsampled_a = FF.interpolate(xa, (224, 224), mode='bilinear', align_corners=False).to('cuda')
            au = torch.sigmoid(unlearnt.layer4[1].conv1.alpha)
            # MI(xa[:, :].mean(dim=(-1, -2)).detach().cpu().numpy(), np.ones(10), random_state=0)
            # MI(xa[:, :].mean(dim=(-1, -2)).detach().cpu().numpy(), np.ones(10), random_state=0)
            # MI(xa[:, :].mean(dim=(-1, -2)).detach().cpu().numpy(), np.ones(10), random_state=0)
            # MI(xa[:, :].mean(dim=(-1, -2)).detach().cpu().numpy(), np.ones(10), random_state=0)
            # au, ag = torch.sigmoid(unlearnt.layer4[1].conv1.alpha), torch.sigmoid(general.layer4[1].conv1.alpha)
            # zu = torch.ones([1, 1, 224, 224], device='cuda')
            # for i in torch.where((1 - au) >.6)[0]:
            #     zu *= upsampled_b[:, i]
            # zu_ = zu - zu.min()
            # zu_ = zu / (zu.max()+1e-10)
            # zg = torch.ones([1, 1, 224, 224], device='cuda')
            # for i in torch.where((1 - ag) >.6)[0]:
            #     zg *= upsampled_a[:, i]
            # zg_ = zg - zg.min()
            # zg_ = zg / (zg.max()+1e-10)

            x=0
        xb = xb.cpu() #torch.nn.functional.relu(xb)

        alpha= unlearnt.layer4[1].conv1.alpha # unlearnt.features[28].alpha
        # alpha = torch.sigmoid(alpha)
        # torch.where()
        if False:
            for i in (alpha).topk(10).indices:
                # test = (upsampled_b[:, i])[(None,) + (...,)] * _T(immagine).cuda()
                # test -= test.min()
                # test /= (test.max() + 1e-12)
                # plt.imshow(test[0].permute(1, 2, 0).detach().cpu().numpy(), cmap='jet')
                # plt.show()

                test = (upsampled_b[:, i].clone())
                # print(test.max(), test.min(), '----')
                # test -= test.min()
                # test /= (test.max() + 1e-12)
                # print(test.max(), test.min())
                plt.imshow(test[0].detach().cpu().numpy())
                plt.show()

            for i in (1-alpha).topk(10).indices:
                test = (upsampled_b[:, i])[(None,) + (...,)] * _T(immagine).cuda()
                test -= test.min()
                test /= (test.max() + 1e-12)
                plt.imshow(test[0].permute(1, 2, 0).detach().cpu().numpy(), cmap='jet')
                plt.show()

                test = (upsampled_b[:, i].clone())
                print(test.max(), test.min(),'----')
                test -= test.min()
                test /= (test.max() + 1e-12)
                print(test.max(), test.min())
                plt.imshow(test[0].detach().cpu().numpy(), cmap='jet')
                plt.show()

            upsum = upsampled_b.sum(dim=1)
            upsum -= upsum.min()
            upsum /= (upsum.max() + 1e-12)
            test = (upsum) * _T(immagine).cuda()
            test -= test.min()
            test /= (test.max() + 1e-12)
            plt.imshow(test[0].permute(1, 2, 0).detach().cpu().numpy(), cmap='jet')
            plt.show()


        # if '29' in ak or True:
        #     try:
        #         path='/homes/spoppi/pycharm_projects/inspecting_twin_models/activations_png/vgg16'
        #         path = os.path.join(path, f'class_{c_to_del}',f'imgidx_{img_idx}', ak)
        #         if not os.path.isdir(path):
        #             os.makedirs(path)
        #         plot_pairs(
        #             xa, xb, 0, ak, idx2=-2,
        #             path=path
        #         )
        #     except Exception as e:
        #         print(e)


        # if 'conv' in ak:
        #     # xa, xb = norm(xa, xb)
        #     val = conv_diff(xa[0], xb[0])
        # else:
        #     val = others_diff(xa, xb)

        # ws[f'A{row}'] = ak
        # row += 1
        # first_row = row
        # ws, row = pyxlutils.write_layer(
        #     val=val,
        #     ws=ws,
        #     st_row=first_row
        #
        # )



# f(AH, BH)

# ws = pyxlutils.color_scale(ws=ws, cols=['M','N','O','P'])
# ws = pyxlutils.adapt_cols_len(ws=ws)
# wb.save(out_name)