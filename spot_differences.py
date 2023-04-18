import os
import torch
import torchvision.models as models
import torch.nn.functional as FF
import torchvision.transforms as transforms
from custom_archs import convert_conv2d_to_alpha
import torchvision.transforms.functional as F
import pandas as pd
import openpyxl as xls
from openpyxl import Workbook
from scipy.stats import pearsonr
from scipy.special import kl_div
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import colors
from torchsummary import summary
from generics.plotutils import plot_pairs

import pyxlutils

size=224
means = [0.485, 0.456, 0.406]
stds = [0.229, 0.224, 0.225]
DeT = transforms.Compose([
    transforms.Normalize(-1 * torch.Tensor(means) / torch.Tensor(stds), 1.0 / torch.Tensor(stds))
])

general = models.vgg16(pretrained=True)

c_to_del = [2] #, 3, 394, 4, 149]

PATH = f"/homes/spoppi/pycharm_projects/inspecting_twin_models/checkpoints/short/vgg16_0.9.pt"
# PATH = f"/homes/spoppi/pycharm_projects/inspecting_twin_models/checkpoints/short/vgg16_0.9_alpha.pt"
# unlearnt = convert_conv2d_to_alpha(models.vgg16(pretrained=False))
unlearnt = models.vgg16(pretrained=False)


unlearnt.load_state_dict(torch.load(PATH))

# summary(unlearnt.cpu(), (3,224,224), device='cpu')

wb = Workbook()
# out_name = f'differences_{"-".join(str(c) for c in c_to_del)}_freeze-before-28_vgg16.xlsx'
out_name = f'differences_vgg16_90_ur100_alpha.xlsx'
ws = wb.active

ws = pyxlutils.create_ws(ws=ws, cname='shark')

def norm(a, b):
    mins = a.flatten(start_dim=-2).min(-1).values.unsqueeze(-1).unsqueeze(-1).clone()
    # b -= b.flatten(start_dim=-2).min(-1).values.unsqueeze(-1).unsqueeze(-1)
    # a_max_nonzero = a.flatten(start_dim=-2).max(-1).values.unsqueeze(-1).unsqueeze(-1)
    bmins = b.flatten(start_dim=-2).min(-1).values.unsqueeze(-1).unsqueeze(-1).clone()
    indices = mins > bmins
    mins[indices] = bmins[indices]

    # maxs = torch.where(maxs != 0, maxs,
    #                    torch.tensor(10e-8).to(device=maxs.device))

    a -= mins
    b -= mins

    maxs = a.flatten(start_dim=-2).max(-1).values.unsqueeze(-1).unsqueeze(-1).clone()
    # b -= b.flatten(start_dim=-2).min(-1).values.unsqueeze(-1).unsqueeze(-1)
    # a_max_nonzero = a.flatten(start_dim=-2).max(-1).values.unsqueeze(-1).unsqueeze(-1)
    bmax = b.flatten(start_dim=-2).max(-1).values.unsqueeze(-1).unsqueeze(-1).clone()
    indices = maxs > bmax
    maxs[indices] = bmax[indices]

    maxs = torch.where(
        maxs != 0, maxs,
        torch.tensor(10e-8).to(device=maxs.device)
    )
    # b_max_nonzero = b.flatten(start_dim=-2).max(-1).values.unsqueeze(-1).unsqueeze(-1)
    # b_max_nonzero = torch.where(b_max_nonzero != 0, b_max_nonzero,
    #                             torch.tensor(10e-8).to(device=b_max_nonzero.device))


    a /= maxs
    b /= maxs

    return a, b


def conv_diff(a,b):
    ret = []
    for i in range(a.shape[0]):
        ret.append([
            [a[i].min(),a[i].max(), a[i].mean(), a[i].var()],
            [b[i].min(),b[i].max(), b[i].mean(), b[i].var()]
        ])

        # kl1 = kl_div(a[i],b[i])
        # kl2 = kl_div(b[i],a[i])
        #
        # ret.append([
        #     [kl1.min(),kl1.max(), kl1.mean(), kl1.var()],
        #     [kl2.min(),kl2.max(), kl2.mean(), kl2.var()]
        # ])

    return ret

def others_diff(a,b):
    ret = []
    ret.append([
        [a.min(),a.max(),a.mean(),a.std()],
        [b.min(),b.max(),b.mean(),b.std()]
    ])

    # kl1 = kl_div(a, b)
    # kl2 = kl_div(b, a)
    #
    # ret.append([
    #     [kl1.min(), kl1.max(), kl1.mean(), kl1.var()],
    #     [kl2.min(), kl2.max(), kl2.mean(), kl2.var()]
    # ])

    return ret

def f(a, b, start_row=1,method='diff', N=True):
    global ws

    ret = dict()
    row = start_row
    for i, (aa, bb) in enumerate(zip(a.named_parameters(), b.named_parameters())):
        # print(aa[0], '\t', (aa[1]-bb[1]).sum())
        xa, xb = aa[1].detach().data, bb[1].detach().data


        if 'alpha' in aa[0]:
            xa = torch.sigmoid(xa)
            xb = torch.sigmoid(xb)
            alpha = xb.clone()
        else:
            xa = torch.nn.functional.relu(xa)
            xb = torch.nn.functional.relu(xb)



        if 'weight' in aa[0] and ('conv' in aa[0] or 'features' in aa[0]):
            # wei_xb = alpha[(...,) + (None,)*3] * xb
            try:
                plot_pairs(xa, xb, idx=0, layer=aa[0], idx2=-1, inc=len(xa))
            except Exception as e:
                print(e)
            # xa, xb = norm(xa, xb)
            val = conv_diff(xa, xb)
        else:
            val = others_diff(xa, xb)

        if 'alpha' in aa[0]:
            ws[f'A{row}'] = aa[0]
            row += 1
            ws, row = pyxlutils.write_layer(
                ws=ws,
                val=val,
                st_row=row
            )

        # sort = {k:v for k,v in sorted(enumerate(val), key=lambda x: x[1], reverse=True)}
        # row=first_row
        # for k,v in sort.items():
        #     ws[f'C{row}'] = k
        #     ws[f'D{row}'] = float(v.detach().data.numpy())
        #     row += 1


        # mxb, sxb = xb.mean(), xb.var()
        #
        # if method == 'diff':
        #     val = torch.abs((xa - xb).sum())
        #
        # elif method == 'inner':
        #     val = torch.inner(xa.flatten(start_dim=-2), xb.flatten(start_dim=-2))
        #
        # elif method == 'stats':
        #
        #     astats = torch.Tensor([mxa, sxa])
        #     bstats = torch.Tensor([mxb, sxb])
        #
        #     val = torch.abs(torch.inner(astats, bstats))
        #
        # ret[i] = {'name': aa[0], 'diffs': val.data.numpy(), 'a_mean': mxa.data.numpy(),
        #           'a_stds': sxa.data.numpy(), 'b_mean': mxb.data.numpy(), 'b_stds': sxb.data.numpy()}

    return ret

diff=f(general,unlearnt,method='diff',N=False,start_row=3)

ws = pyxlutils.color_scale(ws=ws, cols=['M','N','O','P'])
# ws = pyxlutils.color_scale(ws=ws, cols=['C','D','E','F'])
ws = pyxlutils.adapt_cols_len(ws=ws)



wb.save(out_name)

# diff = sorted(diff.items(), key=lambda x:x[1]['diffs'], reverse=True)
# df_ = dict()
# for i, kk in enumerate(diff):
#     for k,v in kk[1].items():
#         if i == 0:
#             df_[k]=[v]
#         else:
#             df_[k].append(v)
#
# df = pd.DataFrame(df_)
# df.to_excel('esempio.xlsx')
